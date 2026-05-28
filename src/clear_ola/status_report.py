"""Build a comprehensive 4-category status report for sharing with the GST
team / client. The four categories:

  Sheet 1 - OTP Required           : NOT_DOWNLOADED rows from partial-items.csv
                                     (Clear's stored GSTN session expired;
                                     needs OTP reconnect in ClearGST UI)
  Sheet 2 - Partial Data Only      : DOWNLOADED_PARTIALLY rows from partial-items
                                     (Clear has incomplete data; confirm with
                                     GST team whether full data should exist)
  Sheet 3 - Downloaded Complete    : `done` rows from manifest (Excel file
                                     written successfully)
  Sheet 4 - No Data Available      : `no_data` rows from manifest (PAN wasn't
                                     yet registered in this FY)

Output:
  state/status-report.xlsx  - 4 sheets + a Summary sheet
  state/status-report.txt   - plain-text summary for chat/email
"""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_status_report(
    *,
    pan_to_business: dict[str, str],
    partials_csv: Path,
    manifest_rows: list[dict],
    xlsx_out: Path,
    txt_out: Path | None = None,
) -> dict:
    """Build the 4-category status report.

    Args:
        pan_to_business: PAN -> business name (from config + user_gstins
            enrichment). Used to enrich sheets 3 & 4 where the manifest doesn't
            store business name.
        partials_csv: path to state/partial-items.csv (may not exist; sheets 1 & 2
            will then be empty).
        manifest_rows: list of dicts from `Manifest.all_rows()`.
        xlsx_out: where to write the Excel.
        txt_out: optional path for the text summary.

    Returns: dict with per-category counts.
    """
    # ---- Read partials (sheets 1 & 2) ----
    otp_rows, partial_rows = _read_partials(partials_csv)

    # ---- Pull done / no_data / failed from manifest (sheets 3 & 4 + summary) ----
    done_rows = [r for r in manifest_rows if r["status"] == "done"]
    no_data_rows = [r for r in manifest_rows if r["status"] == "no_data"]
    failed_rows = [r for r in manifest_rows if r["status"] == "failed"]

    # ---- Write Excel ----
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Default sheet becomes Summary
    summary_ws = wb.active
    summary_ws.title = "0-Summary"

    _write_summary_sheet(
        summary_ws,
        counts={
            "otp_required_pages": len(otp_rows),
            "otp_required_gstins": sum(len(r["gstins"]) for r in otp_rows),
            "partial_data_pages": len(partial_rows),
            "partial_data_gstins": sum(len(r["gstins"]) for r in partial_rows),
            "downloaded": len(done_rows),
            "no_data": len(no_data_rows),
            "failed_other": len(failed_rows),
        },
    )

    _write_otp_sheet(wb.create_sheet("1-OTP Required"),
                     otp_rows, pan_to_business)
    _write_partial_sheet(wb.create_sheet("2-Partial Data Only"),
                         partial_rows, pan_to_business)
    _write_downloaded_sheet(wb.create_sheet("3-Downloaded Complete"),
                            done_rows, pan_to_business)
    _write_no_data_sheet(wb.create_sheet("4-No Data Available"),
                         no_data_rows, pan_to_business)

    wb.save(xlsx_out)

    # ---- Plain-text summary ----
    if txt_out is not None:
        _write_txt_summary(
            txt_out,
            otp_rows=otp_rows,
            partial_rows=partial_rows,
            done_rows=done_rows,
            no_data_rows=no_data_rows,
            failed_rows=failed_rows,
            pan_to_business=pan_to_business,
        )

    return {
        "otp_required_pages": len(otp_rows),
        "otp_required_gstins": sum(len(r["gstins"]) for r in otp_rows),
        "partial_data_pages": len(partial_rows),
        "partial_data_gstins": sum(len(r["gstins"]) for r in partial_rows),
        "downloaded": len(done_rows),
        "no_data": len(no_data_rows),
        "failed_other": len(failed_rows),
    }


# ---------- partials parsing ----------

def _read_partials(partials_csv: Path) -> tuple[list[dict], list[dict]]:
    """Read partials CSV, dedup per (pan, fy, gstin, status), group by (pan, fy).
    Returns (otp_rows, partial_rows) — each row aggregates the GSTINs for one
    (pan, fy) page."""
    if not partials_csv.exists():
        return [], []
    by_key: dict[tuple, dict] = {}  # (pan, fy, gstin, status) -> latest row
    with partials_csv.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = (row["pan"], row["fy"], row["gstin"],
                   (row.get("status") or "").strip())
            by_key[key] = row

    def group(status: str) -> list[dict]:
        grouped: dict[tuple, list[dict]] = defaultdict(list)
        for row in by_key.values():
            if (row.get("status") or "").strip() != status:
                continue
            grouped[(row["pan"], row["fy"])].append(row)
        out = []
        for (pan, fy), items in sorted(grouped.items()):
            biz = items[0].get("business_name", "")
            gstins = sorted({r["gstin"] for r in items})
            states = sorted({r["state_name"] for r in items
                             if r.get("state_name")})
            last_seen = max(r.get("logged_at", "") for r in items)
            out.append({
                "pan": pan,
                "business_name": biz,
                "fy": fy,
                "gstins": gstins,
                "state_names": states,
                "items": items,                    # full rows, for partial sheet detail
                "last_seen": last_seen,
            })
        return out

    return group("NOT_DOWNLOADED"), group("DOWNLOADED_PARTIALLY")


# ---------- sheet writers ----------

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _apply_header(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        longest = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, longest + 2), max_width,
        )


def _write_summary_sheet(ws, counts: dict) -> None:
    ws.append(["clear-ola — Status Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated (UTC): "
               f"{datetime.now(timezone.utc).isoformat(timespec='seconds')}"])
    ws.append([])
    ws.append(["Category", "Count", "Notes"])
    _apply_header(ws, ws.max_row)
    ws.append([
        "1. OTP Required (PAN x FY pages)",
        counts["otp_required_pages"],
        f"{counts['otp_required_gstins']} GSTIN(s) total need OTP reconnect "
        f"in ClearGST UI. See sheet '1-OTP Required'.",
    ])
    ws.append([
        "2. Partial Data Only (PAN x FY pages)",
        counts["partial_data_pages"],
        f"{counts['partial_data_gstins']} GSTIN(s) returned only partial data "
        f"even after force-re-download. See sheet '2-Partial Data Only'. "
        f"Confirm with the GST team whether the missing data should exist.",
    ])
    ws.append([
        "3. Downloaded Complete (PAN x FY)",
        counts["downloaded"],
        "Excel file successfully generated and saved to disk. "
        "See sheet '3-Downloaded Complete'.",
    ])
    ws.append([
        "4. No Data Available (PAN x FY)",
        counts["no_data"],
        "Every underlying GSTIN returned NOT_APPLICABLE — entity wasn't yet "
        "registered for GST during that FY. See sheet '4-No Data Available'.",
    ])
    if counts["failed_other"]:
        ws.append([
            "(Other failures still in manifest)",
            counts["failed_other"],
            "These are failures other than the four categories above (e.g. "
            "earlier network blips before retry was added). Re-run "
            "`download --all` and most should resolve.",
        ])
    _autosize(ws, max_width=80)


def _write_otp_sheet(ws, rows: list[dict], pan_to_business: dict) -> None:
    headers = ["PAN", "Business Name", "FY", "GSTINs Count",
               "GSTINs", "State Names", "Last Seen (UTC)", "ClearGST Navigation"]
    ws.append(headers)
    _apply_header(ws, 1)
    for r in rows:
        biz = r.get("business_name") or pan_to_business.get(r["pan"]) or ""
        ws.append([
            r["pan"], biz, r["fy"], len(r["gstins"]),
            ", ".join(r["gstins"]),
            ", ".join(r["state_names"]),
            r["last_seen"],
            f"ClearGST -> GST -> All Reports -> PAN GSTR-2A -> "
            f"select PAN '{r['pan']}' + FY '{r['fy']}' -> "
            f"click 'Generate OTP to connect GSTINs'",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP_TOP
    _autosize(ws)


def _write_partial_sheet(ws, rows: list[dict], pan_to_business: dict) -> None:
    headers = ["PAN", "Business Name", "FY", "GSTINs Count",
               "GSTINs", "State Names", "Download %s",
               "Periods In Scope", "Last Seen (UTC)", "Suggested Action"]
    ws.append(headers)
    _apply_header(ws, 1)
    for r in rows:
        biz = r.get("business_name") or pan_to_business.get(r["pan"]) or ""
        pct_by_state = []
        periods = set()
        for it in r["items"]:
            sn = it.get("state_name") or "?"
            pct = it.get("download_percentage") or ""
            pct_by_state.append(f"{sn}: {pct}%")
            for p in (it.get("periods_in_scope") or "").split(","):
                if p.strip():
                    periods.add(p.strip())
        ws.append([
            r["pan"], biz, r["fy"], len(r["gstins"]),
            ", ".join(r["gstins"]),
            ", ".join(r["state_names"]),
            "\n".join(pct_by_state),
            ", ".join(sorted(periods)),
            r["last_seen"],
            "Confirm with GST team whether the missing months should exist. "
            "Once confirmed, either re-run normally (Clear may have new data) "
            "or accept the gap with --force-partial.",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP_TOP
    _autosize(ws)


def _write_downloaded_sheet(ws, rows: list[dict], pan_to_business: dict) -> None:
    headers = ["PAN", "Business Name", "FY", "Report Type",
               "File Name", "File Size (bytes)",
               "File Path", "Completed At (UTC)"]
    ws.append(headers)
    _apply_header(ws, 1)
    for r in sorted(rows, key=lambda x: (x["pan"], x["fy"])):
        biz = pan_to_business.get(r["pan"], "")
        file_path = r.get("file_path") or ""
        file_name = Path(file_path).name if file_path else ""
        ws.append([
            r["pan"], biz, r["fy"], r["report_type"],
            file_name, r.get("file_bytes") or "",
            file_path, r.get("completed_at") or "",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP_TOP
    _autosize(ws)


def _write_no_data_sheet(ws, rows: list[dict], pan_to_business: dict) -> None:
    headers = ["PAN", "Business Name", "FY", "Report Type",
               "Reason", "Recorded At (UTC)"]
    ws.append(headers)
    _apply_header(ws, 1)
    for r in sorted(rows, key=lambda x: (x["pan"], x["fy"])):
        biz = pan_to_business.get(r["pan"], "")
        ws.append([
            r["pan"], biz, r["fy"], r["report_type"],
            r.get("error_message") or "All GSTINs returned NOT_APPLICABLE",
            r.get("completed_at") or "",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP_TOP
    _autosize(ws)


def _write_txt_summary(
    txt_out: Path,
    *,
    otp_rows: list[dict],
    partial_rows: list[dict],
    done_rows: list[dict],
    no_data_rows: list[dict],
    failed_rows: list[dict],
    pan_to_business: dict,
) -> None:
    """A plain-text summary suitable for pasting into chat/email."""
    txt_out.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    add = lines.append
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    add(f"clear-ola - Status Report")
    add(f"Generated (UTC): {now}")
    add("")
    total = (len(otp_rows) + len(partial_rows)
             + len(done_rows) + len(no_data_rows) + len(failed_rows))
    add(f"Total (PAN x FY) combinations tracked: {total}")
    add("")
    add(f"  1. OTP Required (PAN x FY pages):     {len(otp_rows):>4}  "
        f"({sum(len(r['gstins']) for r in otp_rows)} GSTIN(s) total)")
    add(f"  2. Partial Data Only (PAN x FY pages):{len(partial_rows):>4}  "
        f"({sum(len(r['gstins']) for r in partial_rows)} GSTIN(s) total)")
    add(f"  3. Downloaded Complete (PAN x FY):    {len(done_rows):>4}")
    add(f"  4. No Data Available (PAN x FY):      {len(no_data_rows):>4}")
    if failed_rows:
        add(f"  (Other failures still in manifest):  {len(failed_rows):>4}")
    add("")
    add("=" * 72)
    add("1. OTP REQUIRED  (need manual OTP reconnect in ClearGST UI)")
    add("=" * 72)
    if not otp_rows:
        add("  (none)")
    else:
        for r in otp_rows:
            biz = r.get("business_name") or pan_to_business.get(r["pan"]) or ""
            add(f"\n  PAN: {r['pan']}  ({biz})")
            add(f"  FY:  {r['fy']}  - {len(r['gstins'])} GSTIN(s)")
            add(f"  States: {', '.join(r['state_names'])}")
    add("")
    add("=" * 72)
    add("2. PARTIAL DATA ONLY  (Clear has incomplete data; please confirm)")
    add("=" * 72)
    if not partial_rows:
        add("  (none)")
    else:
        for r in partial_rows:
            biz = r.get("business_name") or pan_to_business.get(r["pan"]) or ""
            add(f"\n  PAN: {r['pan']}  ({biz})")
            add(f"  FY:  {r['fy']}  - {len(r['gstins'])} GSTIN(s)")
            add(f"  States: {', '.join(r['state_names'])}")
    add("")
    add("=" * 72)
    add(f"3. DOWNLOADED COMPLETE  ({len(done_rows)} file(s) saved to disk)")
    add("=" * 72)
    if done_rows:
        for r in sorted(done_rows, key=lambda x: (x["pan"], x["fy"])):
            biz = pan_to_business.get(r["pan"], "")
            add(f"  {r['pan']}  {r['fy']}  {biz}")
    add("")
    add("=" * 72)
    add(f"4. NO DATA AVAILABLE  ({len(no_data_rows)} combinations)")
    add("=" * 72)
    if no_data_rows:
        for r in sorted(no_data_rows, key=lambda x: (x["pan"], x["fy"])):
            biz = pan_to_business.get(r["pan"], "")
            add(f"  {r['pan']}  {r['fy']}  {biz}")
    add("")
    txt_out.write_text("\n".join(lines), encoding="utf-8")
